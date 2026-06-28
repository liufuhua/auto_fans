from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import Select, func, or_, select
from sqlalchemy.orm import Session

from app.core.exceptions import AppException
from app.models.comment_bank import CommentBankItem
from app.models.device import Device
from app.models.doctor import Doctor, DoctorKeyword
from app.schemas.comment_bank import CommentBankImportResponse, CommentBankItemRead
from app.schemas.common import PageParams, PageResult
from app.services.doctors import get_doctor_or_404


def _to_comment_bank_item_read(
    item: CommentBankItem, doctor_name: str, keyword: str, used_device_name: str | None
) -> CommentBankItemRead:
    return CommentBankItemRead(
        id=item.id,
        doctor_id=item.doctor_id,
        doctor_name=doctor_name,
        keyword_id=item.keyword_id,
        keyword=keyword,
        content=item.content,
        status=item.status,
        used_device_name=used_device_name,
        used_account=item.used_account,
        used_at=item.used_at,
        created_at=item.created_at,
    )


def _comment_bank_base_statement() -> Select[tuple[CommentBankItem, str, str, str | None]]:
    return (
        select(CommentBankItem, Doctor.name, DoctorKeyword.keyword, Device.name)
        .join(Doctor, Doctor.id == CommentBankItem.doctor_id)
        .join(DoctorKeyword, DoctorKeyword.id == CommentBankItem.keyword_id)
        .outerjoin(Device, Device.id == CommentBankItem.used_device_id)
    )


def _apply_comment_bank_filters(
    statement: Select[tuple[CommentBankItem, str, str, str | None]],
    doctor_id: int | None,
    keyword_id: int | None,
    status: str | None,
    keyword: str | None,
) -> Select[tuple[CommentBankItem, str, str, str | None]]:
    if doctor_id:
        statement = statement.where(CommentBankItem.doctor_id == doctor_id)
    if keyword_id:
        statement = statement.where(CommentBankItem.keyword_id == keyword_id)
    if status:
        statement = statement.where(CommentBankItem.status == status)
    if keyword:
        keyword_like = f"%{keyword.strip()}%"
        statement = statement.where(
            or_(
                CommentBankItem.content.like(keyword_like),
                Doctor.name.like(keyword_like),
                DoctorKeyword.keyword.like(keyword_like),
            )
        )
    return statement


def list_comment_bank_items(
    db: Session,
    page_params: PageParams,
    doctor_id: int | None,
    keyword_id: int | None,
    status: str | None,
    keyword: str | None,
) -> PageResult[CommentBankItemRead]:
    statement = _apply_comment_bank_filters(
        _comment_bank_base_statement(), doctor_id, keyword_id, status, keyword
    )
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    rows = db.execute(
        statement.order_by(CommentBankItem.id.desc())
        .offset((page_params.page - 1) * page_params.page_size)
        .limit(page_params.page_size)
    ).all()
    return PageResult(
        items=[
            _to_comment_bank_item_read(item, doctor_name, keyword_text, used_device_name)
            for item, doctor_name, keyword_text, used_device_name in rows
        ],
        total=total,
    )


def delete_comment_bank_item(db: Session, item_id: int) -> None:
    item = db.get(CommentBankItem, item_id)
    if item is None:
        raise AppException("评论词库不存在", code="COMMENT_BANK_ITEM_NOT_FOUND", status_code=404)
    db.delete(item)
    db.commit()


def batch_delete_comment_bank_items(db: Session, item_ids: list[int]) -> int:
    unique_ids = list(dict.fromkeys(item_ids))
    if not unique_ids:
        return 0

    items = db.scalars(select(CommentBankItem).where(CommentBankItem.id.in_(unique_ids))).all()
    for item in items:
        db.delete(item)
    db.commit()
    return len(items)


def _parse_comment_excel(file_bytes: bytes) -> list[tuple[str, str]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise AppException(
            "Excel 文件解析失败", code="EXCEL_PARSE_FAILED", status_code=400
        ) from exc

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    search_word_index = _find_header_index(headers, ["搜索词", "searchWord", "keyword"])
    content_index = _find_header_index(headers, ["评论内容", "content"])
    if search_word_index is None or content_index is None:
        raise AppException(
            "Excel 必须包含“搜索词”和“评论内容”两列", code="EXCEL_HEADER_INVALID", status_code=400
        )

    parsed_rows: list[tuple[str, str]] = []
    for row in rows[1:]:
        search_word = _cell_to_str(row[search_word_index] if search_word_index < len(row) else None)
        content = _cell_to_str(row[content_index] if content_index < len(row) else None)
        if search_word and content:
            parsed_rows.append((search_word, content))
    return parsed_rows


def _find_header_index(headers: list[str], candidates: list[str]) -> int | None:
    for candidate in candidates:
        if candidate in headers:
            return headers.index(candidate)
    return None


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def import_comment_bank_excel(
    db: Session, doctor_id: int, file_bytes: bytes
) -> CommentBankImportResponse:
    get_doctor_or_404(db, doctor_id)
    rows = _parse_comment_excel(file_bytes)
    keywords = db.scalars(select(DoctorKeyword).where(DoctorKeyword.doctor_id == doctor_id)).all()
    keyword_map = {keyword.keyword: keyword for keyword in keywords}

    imported = 0
    skipped = 0
    for search_word, content in rows:
        matched_keyword = keyword_map.get(search_word)
        if matched_keyword is None:
            skipped += 1
            continue

        db.add(
            CommentBankItem(
                doctor_id=doctor_id,
                keyword_id=matched_keyword.id,
                search_word=search_word,
                content=content,
                status="unused",
            )
        )
        imported += 1

    db.commit()
    return CommentBankImportResponse(imported=imported, skipped=skipped)
