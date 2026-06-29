from datetime import date
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import Select, func, or_, select
from sqlalchemy.orm import Session

from app.models.automation_result import AutomationResult
from app.models.comment_recheck import CommentRecheckRecord
from app.models.daily_task import DailyTask
from app.models.device import Device
from app.models.doctor import Doctor, DoctorKeyword
from app.schemas.automation_result import AutomationResultItemRead
from app.schemas.common import PageParams, PageResult
from app.services.comment_recheck import normalize_comment_recheck_status


def _automation_result_base_statement() -> Select[
    tuple[AutomationResult, DailyTask, str, str | None, str, CommentRecheckRecord | None]
]:
    return (
        select(
            AutomationResult,
            DailyTask,
            Doctor.name,
            DoctorKeyword.keyword,
            Device.name,
            CommentRecheckRecord,
        )
        .join(DailyTask, DailyTask.id == AutomationResult.task_id)
        .join(Doctor, Doctor.id == AutomationResult.doctor_id)
        .outerjoin(DoctorKeyword, DoctorKeyword.id == AutomationResult.keyword_id)
        .join(Device, Device.id == AutomationResult.device_id)
        .outerjoin(
            CommentRecheckRecord,
            CommentRecheckRecord.automation_result_id == AutomationResult.id,
        )
        .where(AutomationResult.status.in_(["success", "failed"]))
    )


def _apply_automation_result_filters(
    statement: Select[tuple[AutomationResult, DailyTask, str, str | None, str, CommentRecheckRecord | None]],
    task_id: int | None,
    doctor_id: int | None,
    keyword_id: int | None,
    device_id: int | None,
    status: str | None,
    keyword: str | None,
) -> Select[tuple[AutomationResult, DailyTask, str, str | None, str, CommentRecheckRecord | None]]:
    if task_id:
        statement = statement.where(AutomationResult.task_id == task_id)
    if doctor_id:
        statement = statement.where(AutomationResult.doctor_id == doctor_id)
    if keyword_id:
        statement = statement.where(AutomationResult.keyword_id == keyword_id)
    if device_id:
        statement = statement.where(AutomationResult.device_id == device_id)
    if status:
        statement = statement.where(AutomationResult.status == status)
    if keyword:
        keyword_like = f"%{keyword.strip()}%"
        statement = statement.where(
            or_(
                AutomationResult.comment_content.like(keyword_like),
                AutomationResult.publish_account.like(keyword_like),
                AutomationResult.fail_reason.like(keyword_like),
            )
        )
    return statement


def _to_automation_result_item_read(
    result: AutomationResult,
    task: DailyTask,
    doctor_name: str,
    keyword: str | None,
    device_name: str,
    recheck_record: CommentRecheckRecord | None,
) -> AutomationResultItemRead:
    return AutomationResultItemRead(
        id=result.id,
        task_id=result.task_id,
        task_date=task.task_date,
        doctor_id=result.doctor_id,
        doctor_name=doctor_name,
        keyword_id=result.keyword_id,
        keyword=keyword or "",
        device_id=result.device_id,
        device_name=device_name,
        publish_account=result.publish_account,
        comment_content=result.comment_content,
        video_link=result.video_link,
        status=result.status,
        comment_recheck_status=normalize_comment_recheck_status(
            recheck_record.status if recheck_record else None
        ),
        comment_recheck_fail_reason=recheck_record.fail_reason if recheck_record else None,
        comment_recheck_checked_at=recheck_record.checked_at if recheck_record else None,
        result_summary=result.result_summary,
        fail_reason=result.fail_reason,
        started_at=result.started_at,
        finished_at=result.finished_at,
        screenshot_url=result.screenshot_url,
        log_url=result.log_url,
    )


def list_automation_results(
    db: Session,
    page_params: PageParams,
    task_id: int | None,
    doctor_id: int | None,
    keyword_id: int | None,
    device_id: int | None,
    status: str | None,
    keyword: str | None,
) -> PageResult[AutomationResultItemRead]:
    statement = _apply_automation_result_filters(
        _automation_result_base_statement(),
        task_id,
        doctor_id,
        keyword_id,
        device_id,
        status,
        keyword,
    )
    total = db.scalar(select(func.count()).select_from(statement.subquery())) or 0
    rows = db.execute(
        statement.order_by(AutomationResult.id.desc())
        .offset((page_params.page - 1) * page_params.page_size)
        .limit(page_params.page_size)
    ).all()
    return PageResult(
        items=[
            _to_automation_result_item_read(
                result,
                task,
                doctor_name,
                keyword_text,
                device_name,
                recheck_record,
            )
            for result, task, doctor_name, keyword_text, device_name, recheck_record in rows
        ],
        total=total,
    )


def export_automation_result_summary_by_date_range(
    db: Session, start_date: date, end_date: date
) -> BytesIO:
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    rows = db.execute(
        select(
            DailyTask.task_date,
            Doctor.name,
            Doctor.real_name,
            DoctorKeyword.keyword,
            func.count(AutomationResult.id),
        )
        .join(DailyTask, DailyTask.id == AutomationResult.task_id)
        .join(Doctor, Doctor.id == AutomationResult.doctor_id)
        .outerjoin(DoctorKeyword, DoctorKeyword.id == AutomationResult.keyword_id)
        .where(DailyTask.task_date >= start_date)
        .where(DailyTask.task_date <= end_date)
        .where(AutomationResult.status == "success")
        .group_by(
            DailyTask.task_date,
            Doctor.id,
            Doctor.name,
            Doctor.real_name,
            DoctorKeyword.id,
            DoctorKeyword.keyword,
            AutomationResult.keyword_id,
        )
        .order_by(
            DailyTask.task_date.asc(),
            Doctor.sort_order.asc(),
            Doctor.id.asc(),
            DoctorKeyword.keyword.asc(),
        )
    ).all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["日期", "医生姓名", "关键词", "数量"])

    for row_date, doctor_name, real_name, keyword_text, count in rows:
        export_doctor_name = (real_name or "").strip() or doctor_name
        sheet.append([row_date, export_doctor_name, keyword_text or "", int(count)])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    for row in sheet.iter_rows(
        min_row=1, max_row=max(sheet.max_row, 2), min_col=1, max_col=4
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    column_widths = [14, 22, 22, 10]
    for index, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def save_automation_result_summary_to_desktop(
    db: Session, start_date: date, end_date: date
) -> Path:
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    desktop_dir = _get_desktop_dir()
    filename = f"医生评论统计表_{start_date.isoformat()}_{end_date.isoformat()}.xlsx"
    output_path = desktop_dir / filename
    workbook_stream = export_automation_result_summary_by_date_range(db, start_date, end_date)
    output_path.write_bytes(workbook_stream.getvalue())
    return output_path


def _get_desktop_dir() -> Path:
    home_dir = Path.home()
    for candidate in (home_dir / "Desktop", home_dir / "桌面"):
        if candidate.exists() and candidate.is_dir():
            return candidate
    return home_dir
